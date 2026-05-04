"""
PH-Lending Pro — Excel Export Module (Selective)
Supports exporting specific sheets with optional date filtering.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_connection, rows_to_list, APP_SUPPORT_DIR, BACKUP_DIR


def _style_header(ws, headers, row=1):
    """Apply header styling to a worksheet row."""
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(bottom=Side(style='thin', color='CBD5E1'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def _add_data_rows(ws, data, start_row=2):
    """Add data rows with alternating colors."""
    light_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    data_font = Font(name='Calibri', size=10)

    for i, row_data in enumerate(data):
        row_num = start_row + i
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = data_font
            if i % 2 == 1:
                cell.fill = light_fill


def _add_summary_sheet(wb, conn, date_from=None, date_to=None):
    """Add a Summary sheet with aggregated KPIs."""
    ws = wb.create_sheet("Summary", 0)  # Insert at beginning
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22

    title_font = Font(name='Calibri', bold=True, size=14, color='1E293B')
    head_font  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    fill_dark  = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    fill_blue  = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    data_font  = Font(name='Calibri', size=10)

    ws['A1'] = 'PH-Lending Pro — Export Summary'
    ws['A1'].font = title_font
    ws['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws['B1'].font = Font(name='Calibri', size=10, color='6B7280')

    if date_from or date_to:
        ws['A2'] = f"Date filter: {date_from or 'start'} → {date_to or 'today'}"
        ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='6B7280')

    c = conn.cursor()
    row = 4

    sections = []

    # Client stats
    c.execute("SELECT COUNT(*) FROM clients")
    n_clients = c.fetchone()[0]

    # Loan stats
    c.execute("SELECT COUNT(*), COALESCE(SUM(principal),0), COALESCE(SUM(total_interest),0) FROM loans")
    r = c.fetchone()
    n_loans, total_capital, total_interest_due = r[0], r[1], r[2]

    c.execute("SELECT COUNT(*) FROM loans WHERE status='active'")
    n_active = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM loans WHERE status='paid'")
    n_paid = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM loans WHERE status='defaulted'")
    n_defaulted = c.fetchone()[0]

    # Payment stats
    c.execute("SELECT COUNT(*), COALESCE(SUM(amount),0) FROM payments WHERE voided_at IS NULL")
    rp = c.fetchone()
    n_payments, total_collected = rp[0], rp[1]

    # Commission stats
    c.execute("SELECT COUNT(*), COALESCE(SUM(commission_amount),0) FROM referral_commissions")
    rc = c.fetchone()
    n_commissions, total_commissions = rc[0], rc[1]

    def add_section_header(label):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = head_font
        ws.cell(row=row, column=1).fill = fill_dark
        ws.cell(row=row, column=2).fill = fill_dark
        ws.merge_cells(f'A{row}:B{row}')
        row += 1

    def add_row(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = data_font
        ws.cell(row=row, column=2, value=value).font = Font(name='Calibri', size=10, bold=True)
        row += 1

    add_section_header('📋 CLIENTS')
    add_row('Total clients', n_clients)
    row += 1

    add_section_header('💰 LOANS')
    add_row('Total loans', n_loans)
    add_row('Active loans', n_active)
    add_row('Paid loans', n_paid)
    add_row('Defaulted loans', n_defaulted)
    add_row('Total capital lent (₱)', round(total_capital, 2))
    add_row('Total interest expected (₱)', round(total_interest_due, 2))
    row += 1

    add_section_header('💵 PAYMENTS')
    add_row('Total payment transactions', n_payments)
    add_row('Total amount collected (₱)', round(total_collected, 2))
    add_row('Outstanding (₱)', round(total_capital + total_interest_due - total_collected, 2))
    row += 1

    add_section_header('🤝 COMMISSIONS')
    add_row('Total commissions', n_commissions)
    add_row('Total commission amount (₱)', round(total_commissions, 2))


def export_selective(output_path, sheets, date_from=None, date_to=None):
    """
    Export selected sheets to Excel.

    Args:
        output_path:  Output .xlsx path
        sheets:       List of sheet names to include.
                      Possible values: 'clients', 'loans', 'payments',
                       'amortization', 'penalties', 'commissions'
        date_from:    Optional ISO date string 'YYYY-MM-DD' start filter
        date_to:      Optional ISO date string 'YYYY-MM-DD' end filter

    Returns:
        str: Path to the generated file
    """
    if not sheets:
        sheets = ['clients', 'loans', 'payments', 'amortization', 'penalties', 'commissions']

    conn = get_connection()
    c    = conn.cursor()
    wb   = Workbook()

    # Remove default empty sheet — will be replaced or kept as summary
    default_ws = wb.active
    wb.remove(default_ws)

    # ── Summary sheet always included ─────────────────────────────
    _add_summary_sheet(wb, conn, date_from, date_to)

    # Date helpers
    def date_filter_clause(col, alias=''):
        clauses = []
        if date_from:
            clauses.append(f"{col} >= '{date_from}'")
        if date_to:
            clauses.append(f"{col} <= '{date_to}'")
        return (' AND ' + ' AND '.join(clauses)) if clauses else ''

    # ── Clients ───────────────────────────────────────────────────
    if 'clients' in sheets:
        ws = wb.create_sheet('Clients')
        headers = ["ID", "First Name", "Last Name", "Address", "Contact",
                   "Rating", "Monthly Income", "Referred By", "Notes", "Created"]
        _style_header(ws, headers)
        df = date_filter_clause('created_at')
        c.execute(f"SELECT * FROM clients WHERE 1=1{df} ORDER BY created_at DESC")
        rows = []
        for r in rows_to_list(c.fetchall()):
            rows.append([
                r['id'], r['first_name'], r['last_name'], r['address'],
                r['contact'], r['rating'], r.get('monthly_income', ''),
                r.get('referred_by', ''), r['notes'], r['created_at']
            ])
        _add_data_rows(ws, rows)
        _auto_width(ws)

    # ── Loans ─────────────────────────────────────────────────────
    if 'loans' in sheets:
        ws = wb.create_sheet('Loans')
        headers = ["Loan ID", "Client ID", "Client Name", "Principal (₱)", "Rate %",
                   "Type", "Term (mo)", "Start Date", "Status",
                   "Total Interest (₱)", "Monthly Payment (₱)",
                   "Total Paid (₱)", "Remaining (₱)", "Created"]
        _style_header(ws, headers)
        df = date_filter_clause('l.start_date')
        c.execute(f"""
            SELECT l.*, c.first_name, c.last_name,
                   (SELECT COALESCE(SUM(amount),0) FROM payments WHERE loan_id=l.id AND voided_at IS NULL) as total_paid
            FROM loans l JOIN clients c ON l.client_id = c.id
            WHERE 1=1{df}
            ORDER BY l.created_at DESC
        """)
        rows = []
        for l in rows_to_list(c.fetchall()):
            remaining = round(l['principal'] + l['total_interest'] - l['total_paid'], 2)
            rows.append([
                l['id'], l['client_id'], f"{l['first_name']} {l['last_name']}",
                l['principal'], l['interest_rate'], l['interest_type'],
                l['term_months'], l['start_date'], l['status'],
                l['total_interest'], l['monthly_payment'],
                round(l['total_paid'], 2), remaining, l['created_at']
            ])
        _add_data_rows(ws, rows)
        _auto_width(ws)

    # ── Payments ──────────────────────────────────────────────────
    if 'payments' in sheets:
        ws = wb.create_sheet('Payments')
        headers = ["Payment ID", "Loan ID", "Client ID", "Client Name",
                   "Amount (₱)", "Date", "Method", "Notes", "Created"]
        _style_header(ws, headers)
        df = date_filter_clause('p.payment_date')
        c.execute(f"""
            SELECT p.*, l.client_id, c.first_name, c.last_name
            FROM payments p
            JOIN loans l ON p.loan_id = l.id
            JOIN clients c ON l.client_id = c.id
            WHERE p.voided_at IS NULL{df}
            ORDER BY p.payment_date DESC
        """)
        rows = []
        for p in rows_to_list(c.fetchall()):
            rows.append([
                p['id'], p['loan_id'], p['client_id'],
                f"{p['first_name']} {p['last_name']}",
                p['amount'], p['payment_date'], p['payment_method'],
                p['notes'], p['created_at']
            ])
        _add_data_rows(ws, rows)
        _auto_width(ws)

    # ── Amortization Schedule ─────────────────────────────────────
    if 'amortization' in sheets:
        ws = wb.create_sheet('Amortization Schedule')
        headers = ["Loan ID", "Client Name", "Month #", "Due Date",
                   "Principal (₱)", "Interest (₱)", "Total Due (₱)", "Balance Remaining (₱)"]
        _style_header(ws, headers)
        df = date_filter_clause('a.due_date')
        c.execute(f"""
            SELECT a.*, l.client_id, c.first_name, c.last_name
            FROM amortization_schedule a
            JOIN loans l ON a.loan_id = l.id
            JOIN clients c ON l.client_id = c.id
            WHERE 1=1{df}
            ORDER BY a.loan_id, a.month_number
        """)
        rows = []
        for a in rows_to_list(c.fetchall()):
            rows.append([
                a['loan_id'], f"{a['first_name']} {a['last_name']}",
                a['month_number'], a['due_date'],
                a['principal_portion'], a['interest_portion'],
                a['total_due'], a['balance_remaining']
            ])
        _add_data_rows(ws, rows)
        _auto_width(ws)

    # ── Penalties ─────────────────────────────────────────────────
    if 'penalties' in sheets:
        ws = wb.create_sheet('Penalties')
        headers = ["ID", "Loan ID", "Client ID", "Client Name",
                   "Amount (₱)", "Reason", "Status", "Penalty Date", "Notes"]
        _style_header(ws, headers)
        df = date_filter_clause('p.penalty_date')
        c.execute(f"""
            SELECT p.*, c.first_name, c.last_name
            FROM penalties p JOIN clients c ON p.client_id = c.id
            WHERE 1=1{df}
            ORDER BY p.penalty_date DESC
        """)
        rows = []
        for p in rows_to_list(c.fetchall()):
            rows.append([
                p['id'], p['loan_id'], p['client_id'],
                f"{p['first_name']} {p['last_name']}",
                p['amount'], p['reason'], p['status'],
                p['penalty_date'], p.get('notes', '')
            ])
        _add_data_rows(ws, rows)
        _auto_width(ws)

    # ── Referral Commissions ──────────────────────────────────────
    if 'commissions' in sheets:
        ws = wb.create_sheet('Commissions')
        headers = ["ID", "Referrer ID", "Referrer Name", "Referred ID",
                   "Referred Name", "Loan ID", "Commission (₱)", "Status", "Created"]
        _style_header(ws, headers)
        df = date_filter_clause('rc.created_at')
        c.execute(f"""
            SELECT rc.*,
                   r1.first_name as ref_first, r1.last_name as ref_last,
                   r2.first_name as rfd_first, r2.last_name as rfd_last
            FROM referral_commissions rc
            JOIN clients r1 ON rc.referrer_id = r1.id
            JOIN clients r2 ON rc.referred_id = r2.id
            WHERE 1=1{df}
            ORDER BY rc.created_at DESC
        """)
        rows = []
        for r in rows_to_list(c.fetchall()):
            rows.append([
                r['id'], r['referrer_id'], f"{r['ref_first']} {r['ref_last']}",
                r['referred_id'], f"{r['rfd_first']} {r['rfd_last']}",
                r['loan_id'], r['commission_amount'], r['status'], r['created_at']
            ])
        _add_data_rows(ws, rows)
        _auto_width(ws)

    conn.close()
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    wb.save(output_path)
    return output_path


# Backwards-compatible wrapper
def export_all_to_excel(output_path=None):
    """Export all data (backwards compatible)."""
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(BACKUP_DIR, f"PH-Lending_Export_{timestamp}.xlsx")
    return export_selective(
        output_path,
        ['clients', 'loans', 'payments', 'amortization', 'penalties', 'commissions']
    )
